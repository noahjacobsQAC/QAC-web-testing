# -*- coding: utf-8 -*-

import logging
from openpyxl import Workbook

from src.excel.sheet import Sheet
from src.logger import logger

logger = logging.getLogger(__name__)

# class Xlsx(Sheet):

#     book = Workbook()

#     def __init__(self, filename):
#         Sheet.__init__(self)

class Xlsx:

    book = Workbook()

    def __init__(self, filename):

        logger.debug("initiazling class")
        self.fileName = filename


    def __getattr__(self, attr):
        print(f'{attr}')
        return attr

    def addSheet(self, sheetname, columns):

        self.book.create_sheet(sheetname).append(columns)


    def getSheetNames(self)->list:

        return self.book.sheetnames


    # ? method not exposed currently
    def __getSheetByName(self, sheetname):

        return self.book.get_sheet_by_name(sheetname)


    def deleteSheet(self, sheetname):

        self.book.remove_sheet(self.book.get_sheet_by_name(sheetname))


    def addRows(self, sheetname, rows):

        if sheetname not in self.book.sheetnames:
            print(f'sheet {sheetname} not present in workbook')
            return None

        active = self.book.get_sheet_by_name(sheetname)

        for row in rows:
            logger.info(f'{row}')
            active.append(row)


    # todo add docstring or this method
    def editAtCell(self, sheetname, row, column, value):

        if sheetname not in self.book.sheetnames:
            print(f'sheet {sheetname} not present in workbook')
            return None

        active = self.book.get_sheet_by_name(sheetname)
        active.cell(column=column, row=row, value=f'{value}')


    # ? method not exposed. use generic addRows()
    def __addRow(self, sheetname, row):

        if sheetname not in self.book.sheetnames:
            print(f'sheet {sheetname} not present in workbook')
            return None

        active = self.book.get_sheet_by_name(sheetname)
        active.append(row)


    def editSheetName(self):
        raise NotImplementedError


    def saveWorkbook(self):

        logger.critical("saving workbook")
        # remove empty default sheet
        try:
            self.book.remove_sheet(self.book.get_sheet_by_name("Sheet"))
        except Exception as e:
            logger.warn(f'{e}')

        self.book.save(self.fileName)




